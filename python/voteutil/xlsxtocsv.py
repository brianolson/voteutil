#!/usr/bin/env python3
#
# pip install openpyxl

import csv
import logging
import os

import openpyxl

logger = logging.getLogger(__name__)

def wb_sheet_out(wb, sheetname, outname):
    sheet = wb[sheetname]
    with open(outname, 'wt') as fout:
        writer = csv.writer(fout)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

def newerthan(a, b):
    '''return True if a (source) is newer than b (dest)
    Like `make`, also True if b does not exist
    '''
    if not os.path.exists(b):
        return True
    return os.path.getmtime(a) > os.path.getmtime(b)

def xlsxtocsv(pathin, outdir=''):
    wb = openpyxl.load_workbook(pathin)
    if pathin.endswith('.xlsx'):
        rootname = pathin[:-5]
    elif pathin.endswith('.xls'):
        rootname = pathin[:-4]
    else:
        rootname = pathin
    if outdir:
        outroot = os.path.join(outdir, os.path.basename(rootname))
    else:
        outroot = rootname
    if len(wb.sheetnames) == 1:
        outname = outroot + '.csv'
        if newerthan(pathin, outname):
            logger.info('%s -> %s', pathin, outname)
            wb_sheet_out(wb, wb.sheetnames[0], outname)
        return
    for sn in wb.sheetnames:
        outname = outroot + '_' + sn + '.csv'
        if newerthan(pathin, outname):
            logger.info('%s[%s] -> %s', pathin, sn, outname)
            wb_sheet_out(wb, sn, outname)


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsxfiles', nargs='*')
    ap.add_argument('--verbose', default=False, action='store_true')
    args = ap.parse_args()
    if args.verbose:
        logging.basicConfig(level=logging.DEBUG)
    else:
        logging.basicConfig(level=logging.WARNING)

    for path in args.xlsxfiles:
        xlsxtocsv(path)

if __name__ == '__main__':
    main()
